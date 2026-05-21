import os
import json
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def export_changes_to_excel(changes: List[Dict[str, Any]], output_path: str) -> bool:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "变更列表"

        headers = ["序号", "变更类型", "原内容", "新内容", "位置", "时间"]
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, change in enumerate(changes, 1):
            row = [
                idx,
                change.get('type', ''),
                change.get('old_value', ''),
                change.get('new_value', ''),
                change.get('location', ''),
                change.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        return True
    except Exception as e:
        print(f"导出Excel失败: {e}")
        return False

def export_changes_to_json(changes: List[Dict[str, Any]], output_path: str) -> bool:
    try:
        export_data = {
            "export_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "total_changes": len(changes),
            "changes": changes
        }
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"导出JSON失败: {e}")
        return False

def export_changes_to_txt(changes: List[Dict[str, Any]], output_path: str) -> bool:
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("文档变更列表\n")
            f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"变更总数: {len(changes)}\n")
            f.write("=" * 80 + "\n\n")

            for idx, change in enumerate(changes, 1):
                f.write(f"[{idx}] {change.get('type', '').upper()}\n")
                f.write(f"  位置: {change.get('location', '')}\n")
                f.write(f"  原内容: {change.get('old_value', '')}\n")
                f.write(f"  新内容: {change.get('new_value', '')}\n")
                f.write(f"  时间: {change.get('timestamp', '')}\n")
                f.write("-" * 40 + "\n")
        return True
    except Exception as e:
        print(f"导出TXT失败: {e}")
        return False

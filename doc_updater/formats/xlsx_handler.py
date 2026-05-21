import openpyxl
from openpyxl.cell.cell import Cell
from typing import Dict, List, Any, Tuple
import pandas as pd

class XlsxHandler:
    def __init__(self):
        self.workbook = None
        self.data = []
        self.sheet_names = []

    def read(self, file_path: str) -> bool:
        try:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            self._parse_workbook()
            return True
        except Exception as e:
            print(f"读取Excel文件失败: {e}")
            return False

    def _parse_workbook(self):
        self.data = []
        for sheet_name in self.sheet_names:
            sheet = self.workbook[sheet_name]
            sheet_data = {
                'sheet_name': sheet_name,
                'rows': []
            }
            for row in sheet.iter_rows(values_only=True):
                sheet_data['rows'].append(list(row))
            self.data.append(sheet_data)

    def get_dataframe(self, sheet_index: int = 0) -> pd.DataFrame:
        if sheet_index < len(self.data):
            df = pd.DataFrame(self.data[sheet_index]['rows'])
            return df
        return pd.DataFrame()

    def get_all_dataframes(self) -> Dict[str, pd.DataFrame]:
        result = {}
        for sheet_data in self.data:
            df = pd.DataFrame(sheet_data['rows'])
            result[sheet_data['sheet_name']] = df
        return result

    def get_cell_value(self, sheet_index: int, row: int, col: int) -> Any:
        if sheet_index < len(self.data):
            sheet = self.data[sheet_index]
            if row < len(sheet['rows']):
                if col < len(sheet['rows'][row]):
                    return sheet['rows'][row][col]
        return None

    def set_cell_value(self, sheet_index: int, row: int, col: int, value: Any) -> bool:
        try:
            if self.workbook and sheet_index < len(self.sheet_names):
                sheet = self.workbook[self.sheet_names[sheet_index]]
                sheet.cell(row=row+1, column=col+1, value=value)
                return True
        except Exception as e:
            print(f"设置单元格值失败: {e}")
        return False

    def update_cell(self, sheet_index: int, row: int, col: int, old_value: Any, new_value: Any) -> bool:
        current_value = self.get_cell_value(sheet_index, row, col)
        if str(current_value) == str(old_value):
            return self.set_cell_value(sheet_index, row, col, new_value)
        return False

    def save(self, output_path: str) -> bool:
        try:
            if self.workbook:
                self.workbook.save(output_path)
                return True
        except Exception as e:
            print(f"保存Excel文件失败: {e}")
        return False

    def write_from_template(self, template_path: str, output_path: str) -> bool:
        try:
            wb = openpyxl.load_workbook(template_path)
            for sheet_idx, sheet_data in enumerate(self.data):
                if sheet_idx < len(wb.sheetnames):
                    sheet = wb[wb.sheetnames[sheet_idx]]
                    for row_idx, row in enumerate(sheet_data['rows']):
                        for col_idx, value in enumerate(row):
                            sheet.cell(row=row_idx+1, column=col_idx+1, value=value)
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"从模板写入失败: {e}")
            return False

    def get_all_values(self) -> List[List[Any]]:
        all_values = []
        for sheet_data in self.data:
            for row in sheet_data['rows']:
                all_values.append(row)
        return all_values

    def find_matching_cells(self, search_text: str) -> List[Tuple[int, int, Any]]:
        matches = []
        for sheet_idx, sheet_data in enumerate(self.data):
            for row_idx, row in enumerate(sheet_data['rows']):
                for col_idx, cell_value in enumerate(row):
                    if cell_value and search_text in str(cell_value):
                        matches.append((sheet_idx, row_idx, col_idx, cell_value))
        return matches

    def get_dimensions(self) -> Dict[str, Tuple[int, int]]:
        dims = {}
        for sheet_idx, sheet_data in enumerate(self.data):
            rows = len(sheet_data['rows'])
            max_cols = max(len(row) for row in sheet_data['rows']) if sheet_data['rows'] else 0
            dims[sheet_data['sheet_name']] = (rows, max_cols)
        return dims

    def apply_marking_changes(self, changes: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        applied = []

        for change in changes:
            sheet_name = change.get('sheet_name', '')
            row = change.get('row', 0)
            col_q_new = change.get('column_q_new', 'NA')
            col_r_new = change.get('column_r_new', '')
            col_q_original = change.get('column_q_original', '')
            col_p_value = change.get('column_p_value', '')

            col_status_new = change.get('column_status_new', 'No')
            col_reason_new = change.get('column_reason_new', '')
            col_status_original = change.get('column_status_original', '')
            col_reason_original = change.get('column_reason_original', '')

            if not sheet_name or sheet_name not in self.sheet_names:
                continue

            sheet_idx = self.sheet_names.index(sheet_name)
            sheet = self.workbook[sheet_name]

            if row < 1 or row > sheet.max_row:
                continue

            col_q_index = self._find_column_index_by_header(sheet_idx, 'Q')
            col_r_index = self._find_column_by_name(sheet_idx)
            col_status_index = self._find_column_by_status_header(sheet_idx)
            col_reason_index = self._find_column_by_reason_header(sheet_idx)

            q_applied = False
            r_applied = False
            status_applied = False
            reason_applied = False

            if col_q_index > 0:
                sheet.cell(row=row, column=col_q_index, value=col_q_new)
                q_applied = True

            if col_r_index > 0:
                sheet.cell(row=row, column=col_r_index, value=col_r_new)
                r_applied = True

            if col_status_index > 0:
                sheet.cell(row=row, column=col_status_index, value=col_status_new)
                status_applied = True

            if col_reason_index > 0:
                sheet.cell(row=row, column=col_reason_index, value=col_reason_new)
                reason_applied = True

            if q_applied or r_applied or status_applied or reason_applied:
                applied.append({
                    'sheet_name': sheet_name,
                    'row': row,
                    'column_q_original': col_q_original,
                    'column_q_new': col_q_new,
                    'column_p_value': col_p_value,
                    'column_r_new': col_r_new,
                    'column_status_original': col_status_original,
                    'column_status_new': col_status_new,
                    'column_reason_original': col_reason_original,
                    'column_reason_new': col_reason_new
                })

        return applied

    def _find_column_index_by_header(self, sheet_idx: int, column_letter: str) -> int:
        if sheet_idx >= len(self.data):
            return -1

        letter_upper = column_letter.upper()

        if letter_upper == 'Q':
            for idx, header in enumerate(self.data[sheet_idx].get('rows', [[]])[0]):
                header_str = str(header).strip().upper() if header else ''
                if 'STATUS' in header_str and '供应商' in str(header):
                    return idx + 1
            return 17

        if letter_upper == 'R':
            for idx, header in enumerate(self.data[sheet_idx].get('rows', [[]])[0]):
                header_str = str(header).strip().upper() if header else ''
                if 'COMMENTS' in header_str and '供应商' in str(header):
                    return idx + 1

        header_row = self.data[sheet_idx].get('rows', [[]])[0]
        for idx, header in enumerate(header_row):
            header_str = str(header).strip().upper() if header else ''
            if letter_upper in header_str:
                return idx + 1
        return -1

    def _find_column_by_name(self, sheet_idx: int) -> int:
        return self._find_column_index_by_header(sheet_idx, 'R')

    def _find_column_by_status_header(self, sheet_idx: int) -> int:
        if sheet_idx >= len(self.data):
            return -1

        header_row = self.data[sheet_idx].get('rows', [[]])[0]
        for idx, header in enumerate(header_row):
            header_str = str(header).strip() if header else ''
            header_upper = header_str.upper()
            if '实现状态' in header_str and ('YES/NO' in header_upper or 'Yes/No' in header_str):
                return idx + 1
        return -1

    def _find_column_by_reason_header(self, sheet_idx: int) -> int:
        if sheet_idx >= len(self.data):
            return -1

        header_row = self.data[sheet_idx].get('rows', [[]])[0]
        for idx, header in enumerate(header_row):
            header_str = str(header).strip() if header else ''
            if '不能实现的原因' in header_str:
                return idx + 1
        return -1
